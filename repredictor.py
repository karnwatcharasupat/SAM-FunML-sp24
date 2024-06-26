# -*- coding: utf-8 -*-
"""
Created on Wed Jan 10 01:47:37 2024

@author: Mohammed
"""
import shutil
import sys
from time import sleep

using_colab = False
import numpy as np
import torch
import matplotlib.pyplot as plt
import cv2
from openpyxl import Workbook
import statistics
import os
import random
from IPython import get_ipython
from matplotlib.backend_bases import MouseButton
import time
from PIL import Image
from segment_anything import sam_model_registry, SamPredictor
import matplotlib
import matplotlib.patches as patches


plt.rcParams['keymap.grid'].remove('g')
plt.rcParams['keymap.home'].remove('r')


def show_mask(mask, ax, random_color=False):
    if random_color:
        color = np.concatenate([np.random.random(3), np.array([0.6])], axis=0)
    else:
        color = np.array([30 / 255, 144 / 255, 255 / 255, 0.6])
    h, w = mask.shape[-2:]
    mask_image = mask.reshape(h, w, 1) * color.reshape(1, 1, -1)
    ax.imshow(mask_image)


def show_points(coords, labels, ax, marker_size=50):
    pos_points = coords[labels == 1]
    neg_points = coords[labels == 0]
    ax.scatter(pos_points[:, 0], pos_points[:, 1], color='green', marker='*', s=marker_size, edgecolor='white',
               linewidth=1.25)
    ax.scatter(neg_points[:, 0], neg_points[:, 1], color='red', marker='*', s=marker_size, edgecolor='white',
               linewidth=1.25)
    
def show_box(box_points, ax):
    x1, y1 = box_points[0],box_points[1]
    x2, y2 = box_points[2],box_points[3]

    width = abs(x2 - x1)
    height = abs(y2 - y1)

    rectangle = patches.Rectangle((x1, y1), width, height, alpha=0.5, facecolor='blue')

    ax.add_patch(rectangle)
 

def closetn(node, nodes):
    nodes = np.asarray(nodes)
    deltas = nodes - node
    dist_2 = np.einsum('ij,ij->i', deltas, deltas)
    return np.argmin(dist_2)


sys.path.append("..")


try:
    matplotlib.use('Qt5Agg')
except:
    matplotlib.use('TkAgg')

sam_checkpoint = 'weights/sam_vit_h_4b8939.pth'
model_type = "vit_h"

device = "cuda" if torch.cuda.is_available() else 'cpu'

# %%

first = input("Do you want to load previous work? -y -n -r\n")
while first not in ['y', 'n', 'r']:
    first = input("Chose y or n, Do you want to load previous work? -y -n -r\n")

if first == 'n':
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'slice'
    ws['B1'] = '# green dots of best'
    ws['C1'] = '# red dots of best '
    ws['D1'] = 'SD of green of best '
    ws['E1'] = 'SD of red of best'
    ws['F1'] = 'best score'
    # for i in range(10):
    #     ws[i+'1']='# green dots of '+str(i)
    #     ws[chr(72+i*5)+'1']='# red dots of '+str(i)
    #     ws[chr(73+i*5)+'1']='SD of green of '+str(i)
    #     ws[chr(74+i*5)+'1']='SD of red of '+str(i)
    #     ws[chr(75+i*5)+'1']='score of '+str(i)
    for i in range(9):
        coun = 1
        for col in ws.iter_cols(min_row=1, max_row=1, max_col=12 + i * 5, min_col=7 + i * 5):
            if coun == 1:
                ws[col[0].coordinate] = '# green dots of ' + str(i + 2)
            elif coun == 2:
                ws[col[0].coordinate] = '# red dots of ' + str(i + 2)
            elif coun == 3:
                ws[col[0].coordinate] = 'SD of X of ' + str(i + 2)
            elif coun == 4:
                ws[col[0].coordinate] = 'SD of Y of ' + str(i + 2)
            elif coun == 5:
                ws[col[0].coordinate] = 'score of ' + str(i + 2)
            coun += 1
    name = input("Type your name:\n")

    if not os.path.exists(name):
        os.makedirs(name)
        os.makedirs(os.path.join(name, "masks"))
        os.makedirs(os.path.join(name, "boxes"))
        os.makedirs(os.path.join(name, "points"))
        os.makedirs(os.path.join(name, "sorts"))
        os.makedirs(os.path.join(name, "eachround"))
        os.makedirs(os.path.join(name, "scores"))

    c = 0
    tim = 0
    t = time.time()

else:
    from openpyxl import load_workbook

    name = input("what is your name?\n")
    wb = load_workbook(os.path.join(name, name + ".xlsx"))
    ws = wb.active

    c = len(os.listdir(os.path.join(name, "masks")))

    f = open(os.path.join(name, "time.txt"), 'r')

    tim = f.readline()
    t = time.time()
    f.close()

#### change that later
print(c)
f = False
names = np.load("data/samples.npy", allow_pickle=True)
labels = np.load("data/labels.npy", allow_pickle=True)
n_images = labels.shape[0]

if first in ["y", "n"]:
    indices = range(c, n_images)
else:
    score_path = os.path.join(name, "scores")
    scores = os.listdir(score_path)

    image_ids = np.array([int(sname.replace("score.npy", "")) for sname in scores])
    max_scores = np.array([np.max(np.load(os.path.join(score_path, sname))) for sname in scores])

    above90 = np.sum(max_scores >= 0.9)
    above80 = np.sum((max_scores >= 0.8) & (max_scores < 0.9))
    above70 = np.sum((max_scores >= 0.7) & (max_scores < 0.8))
    above60 = np.sum((max_scores >= 0.6) & (max_scores < 0.7))
    above50 = np.sum((max_scores >= 0.5) & (max_scores < 0.6))
    above40 = np.sum((max_scores >= 0.4) & (max_scores < 0.5))
    above30 = np.sum((max_scores >= 0.3) & (max_scores < 0.4))
    above20 = np.sum((max_scores >= 0.2) & (max_scores < 0.3))
    above10 = np.sum((max_scores >= 0.1) & (max_scores < 0.2))
    above00 = np.sum((max_scores >= 0.0) & (max_scores < 0.1))


    print(f"Total number of images: {n_images}")
    print(f"Scores distribution: \n"
            f">= 90: {above90}\n",
            f"80-90: {above80}\n",
            f"70-80: {above70}\n",
            f"60-70: {above60}\n",
            f"50-60: {above50}\n",
            f"40-50: {above40}\n",
            f"30-40: {above30}\n",
            f"20-30: {above20}\n",
            f"10-20: {above10}\n",
            f" 0-10: {above00}\n")



    threshold = input("Enter the upper threshold for the scores: ")
    threshold = float(threshold)

    lthreshold = input("Enter the lower threshold for the scores: ")
    lthreshold = float(lthreshold)

    print(f"Threshold: {lthreshold} - {threshold}")
    print(f"Number of images with score within threshold: {np.sum((max_scores >= lthreshold) & (max_scores < threshold))}")

    redo_indices = image_ids[(max_scores >= lthreshold) & (max_scores < threshold)]
    indices = np.arange(c, n_images)

    indices = np.concatenate([indices, redo_indices])

    # nbak = 0
    # while os.path.exists(os.path.join(os.getcwd(),  f"{name}.bak{nbak}")):
    #     nbak += 1
    #
    # shutil.copy(os.path.join(os.getcwd(), name), os.path.join(os.getcwd(),  f"{name}.bak{nbak}"))


sam = sam_model_registry[model_type](checkpoint=sam_checkpoint)
sam.to(device=device)

predictor = SamPredictor(sam)


## start looping through samples: 
for c in indices:
    if f:
        break
    msk = []  # masks for each samples

    gp = []  # green points
    rp = []  # red points

    boxp = [] # prompting boxes

    image = names[c]  # samples c
    ws['A' + str(c + 2)] = str(c)  # samples name on excel
    if len(image.shape) == 2:
        image = cv2.cvtColor((np.array(((image + 1) / 2) * 255, dtype='uint8')), cv2.COLOR_GRAY2RGB)
    label = labels[c]  # GT for sample c
    rmv = False
    mask = 0
    # image=np.array(((image+1)/2)*255,dtype='uint8')
    print(f"Setting image {c}")
    t0 = time.time()
    predictor.set_image(image)
    t1 = time.time()
    print("Image set, time:", t1 - t0)
    inc = ""
    co = 0
    bs = 0
    score = []
    round = [0, 0]
    stdx = []
    stdy = []
    ng = []
    nr = []
    green = []
    red = []
    greenx = []

    redx = []
    greeny = []
    redy = []

    box = []
    box_x =[]
    box_y =[]

    label = label == 1

    while inc != "y":
        s = 0  # this is for the score
        count = 1  # to count the score max
        lessfive = 0
        current_color = 'green'
        current_shape = 'dot'
        is_caching = False
        is_merging = False
        old_mask = None
        # get_ipython().run_line_magic('matplotlib', 'qt')
        fig, ax = plt.subplots(1, 3, figsize=(15, 7))

        if green and red:
            ax[0].plot(greenx, greeny, 'go', markersize=5)
            ax[1].plot(greenx, greeny, 'go', markersize=5)
            ax[0].plot(redx, redy, 'ro', markersize=5)
            ax[1].plot(redx, redy, 'ro', markersize=5)
            plt.draw()

        if len(box_x)==2:
            ax[0].plot(box_x[0], box_y[0], 'bo', markersize=5)
            ax[1].plot(box_x[0], box_y[1], 'bo', markersize=5)
          
            plt.draw()

        def onclose(event):
            fig.canvas.stop_event_loop()
            fig.canvas.mpl_disconnect(cid)

    

        def onclick(event):
            global count
            global green
            global red
            global greenx
            global redx
            global greeny
            global redy
            global box
            global box_x
            global box_y
            global label
            global mask
            global lessfive
            global old_mask
            if event.xdata is not None and event.ydata is not None:

                x, y = int(event.xdata), int(event.ydata)
                print(not x)
                print(not y)
                # if not x or not y:
                #     inc=input("do you wish to continue?")
                #     f=True

                if event.button is MouseButton.LEFT:
                    if current_shape == 'dot':
                        if current_color == 'green':

                            green.append((x, y))
                            greenx.append(x)

                            greeny.append(y)
                            ax[0].plot(x, y, 'go', markersize=5)
                            ax[1].plot(x, y, 'go', markersize=5)
                            plt.draw()


                        elif current_color == 'red':
                            red.append((x, y))
                            redx.append(x)

                            redy.append(y)
                            ax[0].plot(x, y, 'ro', markersize=5)
                            ax[1].plot(x, y, 'ro', markersize=5)
                            plt.draw()

                    elif current_shape =='box':
                        box.append((x,y))
                        box_x.append(x)

                        box_y.append(y)
                        ax[0].plot(x, y, 'bo', markersize=5)
                        ax[1].plot(x, y, 'bo', markersize=5)
                        print(box)
                        print(len(box_x))
                        plt.draw()



                elif event.button is MouseButton.RIGHT:

                    if not green and not red:
                        print("no points to delete")
                    elif green:
                        print(current_color)
                        if current_color == 'green':
                            # print("g",len(green))

                            indx = closetn((x, y), green)
                            print(indx)
                            for line in ax[0].lines:
                                if len(line.get_xdata()) > 0:
                                    if line.get_xdata()[0] == green[indx][0] and line.get_ydata()[0] == green[indx][1]:
                                        # print("Here1")
                                        line.set_data([], [])
                                        break
                            for line in ax[1].lines:
                                if len(line.get_xdata()) > 0:
                                    if line.get_xdata()[0] == green[indx][0] and line.get_ydata()[0] == green[indx][1]:
                                        # print("Here2")
                                        line.set_data([], [])
                                        break
                            del green[indx]
                            del greenx[indx]

                            del greeny[indx]

                            # ax[0].plot(x, y, 'go', markersize=5)
                            # ax[1].plot(x, y, 'go', markersize=5)

                            plt.draw()
                        elif red:
                            print("delete red")
                            print(current_color)
                            indx = closetn((x, y), red)
                            print(indx)

                            for line in ax[0].lines:
                                if len(line.get_xdata()) > 0:
                                    print()
                                    if line.get_xdata()[0] == red[indx][0] and line.get_ydata()[0] == red[indx][1]:
                                        line.set_data([], [])
                                        break
                            for line in ax[1].lines:
                                if len(line.get_xdata()) > 0:
                                    if line.get_xdata()[0] == red[indx][0] and line.get_ydata()[0] == red[indx][1]:
                                        line.set_data([], [])
                                        break
                            # ax[0].plot(x, y, 'ro', markersize=5)
                            # ax[1].plot(x, y, 'ro', markersize=5)
                            # ax[0].set_offsets(red)
                            # a.set_offsets(red)
                            del red[indx]
                            del redx[indx]

                            del redy[indx]
                            plt.draw()

                if (green and red) or len(box_x)==2:
                    global s
                    print("green:", green)
                    print("red:", red)

                    input_point = np.concatenate((green, red))
                    input_label = np.concatenate(([1] * len(green), [0] * len(red)))
                    if current_shape=='dot':
                        masks, scores, logits = predictor.predict(
                        point_coords=input_point,
                        point_labels=input_label,
                        multimask_output=True,
                        )
                    elif current_shape == 'box':
                        box_points = np.array([box_x[0],box_y[0],box_x[1],box_y[1]])

                        masks, scores, logits = predictor.predict(
                        box=box_points,
                        multimask_output=True,
                        )

                    mask = masks[0]

                    # get_ipython().run_line_magic('matplotlib', 'inline')
                    ax[2].clear()
                    ax[2].imshow(image)

                    if is_merging:
                        if old_mask is not None:
                            mask = mask | old_mask


                    show_mask(mask, ax[2])

                    if is_caching:
                        old_mask = mask
                    else:
                        old_mask = None


                    intersection = (mask & label).sum()
                    union = (mask | label).sum()
                    if intersection == 0:
                        s = 0
                    else:
                        s = intersection / union
                    # ws[chr(68)+str(c+2)]=str(bs) # start at cell D(c)

                    print("IOU:", s)
                    if current_shape=='dot':
                        show_points(input_point, input_label, ax[2])
                    elif current_shape=='box':
                        show_box(box_points, ax[2])
                        # Sav box points
                        boxp.append(np.multiply(box_points,1))
                        green = []
                        red = []
                        greenx = []
                        redx = []
                        greeny = []
                        redy = []
                        box = []
                        box_x =[]
                        box_y = []
                        for line in ax[0].lines:
                            line.set_data([], [])
                        for line in ax[1].lines:
                            line.set_data([], [])
                    msg = ""

                    
                    if len(score[round[0]:]) == 0:
                        maxx = 0
                    else:
                        maxx = max(score[round[0]:])
                        print("maxx", maxx)
                    score.append(s)
                    if current_shape=='dot':
                        gp.append(np.multiply(green, 1))
                        rp.append(np.multiply(red, 1))
                        ng.append(len(greenx))
                        nr.append(len(redx))

                        grx = np.concatenate([greenx, redx])
                        gry = np.concatenate([greeny, redy])

                        stdx.append(statistics.pstdev(grx.astype(int).tolist()))
                        stdy.append(statistics.pstdev(gry.astype(int).tolist()))
                    print("up count", count)
                    if maxx >= s:
                        print("inside", count)
                        if count >= 10:

                            # msg="\nNo better score is achieved in the last 5 attempts. Start round 2 from scra\nThe best score ("+str(round(max(score),2))+") is saved"
                            lessfive += 1
                        else:

                            count += 1
                    elif maxx < s:

                        count = 1
                    if lessfive == 1:
                        maxx = 0
                        count = 1
                        round[0] = len(np.array(score))
                        msg = " (round 2) "
                    plt.title(f"Score: {(intersection / union):.3f}" + msg, fontsize=13)
                    ## saving masks, scores, points and other stats: 
                    msk.append(np.multiply(mask, 5))
                    print("less than best score", lessfive)
                    print("scores:", score)
                    if lessfive == 1:
                        lessfive += 1
                        for line in ax[0].lines:
                            line.set_data([], [])
                        for line in ax[1].lines:
                            line.set_data([], [])
                        green = []
                        red = []
                        greenx = []
                        redx = []
                        greeny = []
                        redy = []
                        plt.draw()
                        ax[2].clear()
                        ax[2].imshow(image)
                        show_mask(mask, ax[2])
                        count = 1
                        print("below count", count)
                        plt.title("No better score is achieved in the last 5 attempts. Start round 2 from scratch")
                    elif lessfive == 3:
                        round[1] = len(score) - round[0]
                        print(
                            "The window closed because you did not achieve a better score after 5 consecutive clicks in the 2nd round")
                        plt.close()


        # Create a function to toggle between green and red dots
        def toggle(event):
            global green
            global red
            global greenx
            global redx
            global greeny
            global redy
            global box
            global box_x
            global box_y
            global current_color
            global current_shape
            global is_merging
            global is_caching
            global count
            if event.key == 'g':
                current_color = 'green'
                print("Switched to GREEN dot mode.")

            elif event.key == 'r':
                current_color = 'red'
                print("Switched to RED dot mode.")

            elif event.key =='b':
                current_shape = 'box'
                print("Switched to box mode.")

            elif event.key =='d':
                current_shape = 'dot'
                print("Switched to dot mode.")

            elif event.key == 'm':
                is_merging = True
                print("Merging mode:", is_merging)

            elif event.key == 'u':
                is_merging = False
                print("Merging mode:", is_merging)

            elif event.key == 'c':
                is_caching = True
                print("Caching mode:", is_caching)

            elif event.key == 'x':
                is_caching = False
                print("Caching mode:", is_caching)


            elif event.key == ' ':
                for line in ax[0].lines:
                    line.set_data([], [])
                for line in ax[1].lines:
                    line.set_data([], [])
                green = []
                red = []
                greenx = []
                redx = []
                greeny = []
                redy = []
                box = []
                box_x = []
                box_y = []
                plt.draw()
                ax[2].clear()
                ax[2].imshow(image)
                show_mask(mask, ax[2])
                count = 1
                print("below count", count)



        # Create a figure and display the image

        a = ax[0].plot()
        b = ax[1].plot()
        ax[0].imshow(image)
        ax[1].imshow(label)
        # Connect mouse click and keyboard key events

        fig.canvas.mpl_connect('button_press_event', onclick)
        # fig.canvas.start_event_loop(timeout=-5)
        fig.canvas.mpl_connect('key_press_event', toggle)
        fig.canvas.mpl_connect('key_press_event', toggle)
        # fig.canvas.start_event_loop(timeout=-5)
        # Display the plot

        cid = fig.canvas.mpl_connect('close_event', onclose)
        fig.show()  # this call does not block on my system
        fig.canvas.start_event_loop()  # block here until window closed

        inc = "y"
        print(inc)

    indx = np.argsort(-np.array(score))
    sscore = np.array(score)[indx]
    if current_shape=='dot':
        sng = np.array(ng)[indx]
        snr = np.array(nr)[indx]
        sstdx = np.array(stdx)[indx]
        sstdy = np.array(stdy)[indx]
        for i in range(len(score)):
            coun = 1
            for col in ws.iter_cols(min_row=c + 2, max_row=c + 2, max_col=6 + i * 5, min_col=2 + i * 5):
                if coun == 1:
                    ws[col[0].coordinate] = sng[i]
                elif coun == 2:
                    ws[col[0].coordinate] = snr[i]
                elif coun == 3:
                    ws[col[0].coordinate] = sstdx[i]
                elif coun == 4:
                    ws[col[0].coordinate] = sstdy[i]
                elif coun == 5:
                    ws[col[0].coordinate] = sscore[i]
                coun += 1
    np.save(os.path.join(name, "points", str(c) + "_green"), np.array(gp, dtype=object))
    np.save(os.path.join(name, "points", str(c) + "_red"), np.array(rp, dtype=object))
    np.save(os.path.join(name, "boxes", str(c) + "_box"), np.array(boxp, dtype=object))
    np.save(os.path.join(name, "masks", str(c) + "_mask"), np.array(msk))
    np.save(os.path.join(name, "sorts", str(c) + "_sort"), indx)
    np.save(os.path.join(name, "scores", str(c) + "score"), score)
    np.save(os.path.join(name, "eachround", str(c) + "_"), round)

    # c += 1
    contin = input("Do you want to stop? Press `n` if you do not want to continue, and anything otherwise ")
    if contin == 'n':
        wb.save(os.path.join(name, name + '.xlsx'))
        f = True
        file = open(os.path.join(name, "time.txt"), 'w')
        file.write(str(float(tim) + (time.time() - t)))
        file.close()
    # print("Sample:", c)
wb.save(os.path.join(name, name + '.xlsx'))
